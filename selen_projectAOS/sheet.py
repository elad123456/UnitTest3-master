import openpyxl
from openpyxl import Workbook
import json

class AOS_Sheet:

    # define the values in XL file
    # in each row are the kind of the values and in each column the number of the test
    def __init__(self):
        self.workbook = Workbook()
        self.workbook=openpyxl.load_workbook("Tests_AOS.xlsx")
        self.sheet = self.workbook.active

        self.sheet.title="Tests_AOS"
        self.sheet["A2"]='Product1'
        self.sheet["A6"]='Product2'
        self.sheet["A10"]='Product3'
        self.sheet["A14"]='AOS Existing Account'
        self.sheet["A16"]='AOS New Account'
        self.sheet["A19"]='SafePay User'
        self.sheet["A21"]='MasterCredit'
        self.sheet["A26"]='Test Result'


        self.sheet["B2"] = "Category"
        self.sheet["B3"] = "ID"
        self.sheet["B4"] = "Quantity"
        self.sheet["B5"] = "Color"
        self.sheet["B6"] = "Category"
        self.sheet["B7"] = "ID"
        self.sheet["B8"] = "Quantity"
        self.sheet["B9"] = "Color"
        self.sheet["B10"] = "Category"
        self.sheet["B11"] = "ID"
        self.sheet["B12"] = "Quantity"
        self.sheet["B13"] = "Color"
        self.sheet["B14"] = "username"
        self.sheet["B15"] = "password"
        self.sheet["B16"] = "username"
        self.sheet["B17"] = "mail"
        self.sheet["B18"] = "password"
        self.sheet["B19"] = "username"
        self.sheet["B20"] = "password"
        self.sheet["B21"] = "Card"
        self.sheet["B22"] = "CVV"
        self.sheet["B23"] = "Month"
        self.sheet["B24"] = "Year"
        self.sheet["B25"] = "Name"

        self.sheet["C1"] = "Test 1"
        self.sheet["C2"] = "speakers"
        self.sheet["C3"] = 19
        self.sheet["C4"] = 3
        self.sheet["C5"] = "GRAY"
        self.sheet["C6"] = "tablets"
        self.sheet["C7"] = 18
        self.sheet["C8"] = 4
        self.sheet["C9"] = "BLACK"
        self.sheet["C10"] = "laptops"
        self.sheet["C11"] = 7
        self.sheet["C12"] = 2
        self.sheet["C13"] = "GRAY"

        self.sheet["D1"] = "Test 2"
        self.sheet["D2"] = "headphones"
        self.sheet["D3"] = 12
        self.sheet["D4"] = 2
        self.sheet["D5"] = "PURPLE"
        self.sheet["D6"] = "speakers"
        self.sheet["D7"] = 20
        self.sheet["D8"] = 1
        self.sheet["D9"] = "BLACK"
        self.sheet["D10"] = "mice"
        self.sheet["D11"] = 30
        self.sheet["D12"] = 3
        self.sheet["D13"] = "BLUE"

        self.sheet["E1"] = "Test 3"
        self.sheet["E2"] = "mice"
        self.sheet["E3"] = 31
        self.sheet["E4"] = 6
        self.sheet["E5"] = "BLUE"
        self.sheet["E6"] = "headphones"
        self.sheet["E7"] = 15
        self.sheet["E8"] = 2
        self.sheet["E9"] = "BLACK"
        self.sheet["E10"] = "laptops"
        self.sheet["E11"] = 9
        self.sheet["E12"] = 3
        self.sheet["E13"] = "BLACK"

        self.sheet["F1"] = "Test 4"
        self.sheet["F2"] = "mice"
        self.sheet["F3"] = 34
        self.sheet["F4"] = 1
        self.sheet["F5"] = "GRAY"
        self.sheet["F6"] = "speakers"
        self.sheet["F7"] = 22
        self.sheet["F8"] = 2
        self.sheet["F9"] = "YELLOW"
        self.sheet["F10"] = "headphones"
        self.sheet["F11"] = 14
        self.sheet["F12"] = 3
        self.sheet["F13"] = "GRAY"


        self.sheet["G1"] = "Test 5"
        self.sheet["G2"] = "laptops"
        self.sheet["G3"] = 9
        self.sheet["G4"] = 2
        self.sheet["G5"] = "WHITE"
        self.sheet["G6"] = "speakers"
        self.sheet["G7"] = 23
        self.sheet["G8"] = 7
        self.sheet["G9"] = "BLACK"
        self.sheet["G10"] = "tablets"
        self.sheet["G11"] = 18
        self.sheet["G12"] = 1
        self.sheet["G13"] = "BLACK"

        self.sheet["H1"] = "Test 6"
        self.sheet["H2"] = "mice"
        self.sheet["H3"] = 30
        self.sheet["H4"] = 6
        self.sheet["H5"] = "PURPLE"
        self.sheet["H6"] = "speakers"
        self.sheet["H7"] = 25
        self.sheet["H8"] = 5
        self.sheet["H9"] = "TURQUOISE"
        self.sheet["H10"] = "laptops"
        self.sheet["H11"] = 7
        self.sheet["H12"] = 8
        self.sheet["H13"] = "GRAY"

        self.sheet["I1"] = "Test 7"
        self.sheet["I2"] = "tablets"
        self.sheet["I3"] = 18
        self.sheet["I4"] = 4
        self.sheet["I5"] = "BLACK"
        self.sheet["I6"] = "mice"
        self.sheet["I7"] = 31
        self.sheet["I8"] = 1
        self.sheet["I9"] = "BLACK"
        self.sheet["I10"] = "laptops"
        self.sheet["I11"] = 9
        self.sheet["I12"] = 3
        self.sheet["I13"] = "BLACK"


        self.sheet["J1"] = "Test 8"
        self.sheet["J2"] = "speakers"
        self.sheet["J3"] = 20
        self.sheet["J4"] = 4
        self.sheet["J5"] = "BLACK"
        self.sheet["J6"] = "mice"
        self.sheet["J7"] = 30
        self.sheet["J8"] = 3
        self.sheet["J9"] = "BLUE"
        self.sheet["J10"] = "headphones"
        self.sheet["J11"] = 15
        self.sheet["J12"] = 1
        self.sheet["J13"] = "BLACK"
        self.sheet["J16"] = "el123473"
        self.sheet["J17"] = "el12l3@gmail.com"
        self.sheet["J18"] = "El23416"
        self.sheet["J19"] = "el12345"
        self.sheet["J20"] = "El23456"

        self.sheet["K1"] = "Test 9"
        self.sheet["K2"] = "mice"
        self.sheet["K3"] = 31
        self.sheet["K4"] = 1
        self.sheet["K5"] = "BLACK"
        self.sheet["K6"] = "headphones"
        self.sheet["K7"] = 15
        self.sheet["K8"] = 1
        self.sheet["K9"] = "BLACK"
        self.sheet["K10"] = "laptops"
        self.sheet["K11"] = 9
        self.sheet["K12"] = 2
        self.sheet["K13"] = "WHITE"
        self.sheet["K14"] = "elad1234"
        self.sheet["K15"] = "Thbyrby145"
        self.sheet["K21"] = "123456789123"
        self.sheet["K22"] = "123"
        self.sheet["K23"] = "2"
        self.sheet["K24"] = "3"
        self.sheet["K25"] = "elad-gal"

        self.sheet["L1"] = "Test 10"
        self.sheet["L14"] = "elad1234"
        self.sheet["L15"] = "Thbyrby145"

        self.workbook.save(filename="Tests_AOS.xlsx")
        self.workbook.close()

    # print all the products details in JSON format
    def xl_to_json_products(self):
        tests={}
        subjects=[self.sheet["A2"].value,self.sheet["A6"].value,self.sheet["A10"].value]
        products_details={}
        for column in self.sheet.iter_cols(1,11,1,13,True):
             if column[0] is None or column[1] is None:
                continue
             else:
                 test_number=column[0]
                 x=1
                 for i in subjects:

                    detail={
                     "Category":column[x],
                     "Product ID": column[x+1],
                     "Quantity": column[x+2],
                     "Color": column[x+3]
                    }
                    x+=4
                    products_details[i]=detail
                 tests[test_number]=products_details

        print(json.dumps(tests,indent=4))

    # return the value of the category from XL file
    def get_Category(self,test_number,product_number):
        tests=["C","D","E","F","G","H","I","J","K"]
        if product_number==1:
            category=self.sheet[f"{tests[test_number-1]+'2'}"].value
            return category
        if product_number==2:
            category=self.sheet[f"{tests[test_number-1]+'6'}"].value
            return category
        if product_number==3:
            category=self.sheet[f"{tests[test_number-1]+'10'}"].value
            return category

    # return the id of the product from XL file
    def get_Product_ID(self,test_number,product_number):
        tests=["C","D","E","F","G","H","I","J","K"]
        if product_number == 1:
            id = self.sheet[f"{tests[test_number-1] + '3'}"].value
            return id
        if product_number == 2:
            id = self.sheet[f"{tests[test_number-1] + '7'}"].value
            return id
        if product_number == 3:
            id = self.sheet[f"{tests[test_number-1] + '11'}"].value
            return id

    # return the quantity of the product from XL file
    def get_Quantity(self,test_number,product_number):
        tests=["C","D","E","F","G","H","I","J","K"]
        if product_number == 1:
            quantity = self.sheet[f"{tests[test_number-1] + '4'}"].value
            return quantity
        if product_number == 2:
            quantity = self.sheet[f"{tests[test_number-1] + '8'}"].value
            return quantity
        if product_number == 3:
            quantity = self.sheet[f"{tests[test_number-1] + '12'}"].value
            return quantity

    # return the value of the color from XL file
    def get_Color(self,test_number,product_number):
        tests=["C","D","E","F","G","H","I","J","K","L"]
        if product_number == 1:
            color = self.sheet[f"{tests[test_number-1] + '5'}"].value
            return color
        if product_number == 2:
            color = self.sheet[f"{tests[test_number-1] + '9'}"].value
            return color
        if product_number == 3:
            color = self.sheet[f"{tests[test_number-1] + '13'}"].value
            return color

    # add to the XL file V if it passed and X if it failed
    def add_pass_or_fail(self,test_number,boo):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J","K","L"]
        if boo:
            self.sheet[f"{tests[test_number-1] + '26'}"]="V"
        else:
            self.sheet[f"{tests[test_number - 1] + '26'}"] = "X"
        self.workbook.save(filename="Tests_AOS.xlsx")

    # return the username of exist user
    def get_exist_username(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        username = self.sheet[f"{tests[test_number - 1] + '14'}"].value
        return username

    # return the password of exist user
    def get_exist_password(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        password = self.sheet[f"{tests[test_number - 1] + '15'}"].value
        return password

    # return the username of new user
    def get_new_username(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        username = self.sheet[f"{tests[test_number - 1] + '16'}"].value
        return username

    # return the password of new user
    def get_new_password(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        password = self.sheet[f"{tests[test_number - 1] + '18'}"].value
        return password

    # return the the mail of new user
    def get_new_mail(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        mail = self.sheet[f"{tests[test_number - 1] + '17'}"].value
        return mail

    # return the pay-pal username of specific user
    def get_new_paypal_username(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        username = self.sheet[f"{tests[test_number - 1] + '19'}"].value
        return username

    # return the pay-pal password of specific user
    def get_new_paypal_password(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        password = self.sheet[f"{tests[test_number - 1] + '20'}"].value
        return password

    # return the card-number of specific user
    def get_card_number(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        card_number = self.sheet[f"{tests[test_number - 1] + '21'}"].value
        return card_number

    # return the CVV of specific user
    def get_CVV(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        cvv = self.sheet[f"{tests[test_number - 1] + '22'}"].value
        return cvv

    # return the holder_name of specific user
    def get_holder_name(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        holder_name = self.sheet[f"{tests[test_number - 1] + '25'}"].value
        return holder_name

    # return the MM of specific user
    def get_MM(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        MM = self.sheet[f"{tests[test_number - 1] + '23'}"].value
        return MM

    # return the YYYY of specific user
    def get_YYYY(self,test_number):
        tests = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
        YYYY = self.sheet[f"{tests[test_number - 1] + '24'}"].value
        return YYYY







if __name__=="__main__":
    x=AOS_Sheet()
    x.xl_to_json()
    x.get_Category(3,1)
    x.get_Quantity(3,1)
    x.get_Color(3,1)
    x.get_Product_ID(3,1)