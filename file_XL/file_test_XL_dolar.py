from openpyxl import *
workbook=load_workbook(filename="test.xlsx")
sheet=workbook.active
# for row in sheet.iter_rows(min_row=1,max_row=3,min_col=1,max_col=2,values_only=True):
#     print(row)

for row in sheet.iter_rows():
    print(row)