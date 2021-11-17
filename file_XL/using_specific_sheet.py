from openpyxl import load_workbook
workbook= load_workbook(filename="hello_world.xlsx")
# print the name of the page in this XL file
print(workbook.sheetnames)
sheet = workbook.active
# her i can write in this specific sheet
print(sheet["A1"].value)
sheet["A1"]="elad"
sheet["A3"]="ratner"
sheet["B3"]="LIAV"
print(sheet["A3"].value)
row,col=1,1
print(sheet.cell(row,col).value)
workbook.save(filename="hello_world.xlsx")



